"""
資安新聞收集器 - 核心模組
============================
自動收集多個資安新聞來源的 RSS Feed，
提供分類、搜尋、篩選功能，並可匯出 Excel/CSV/JSON。
"""

import feedparser
import requests
from bs4 import BeautifulSoup
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Optional
from datetime import datetime, timedelta
import pandas as pd
import hashlib
import re
import json
import os
from concurrent.futures import ThreadPoolExecutor, as_completed


@dataclass
class NewsItem:
    """新聞項目資料結構"""
    title: str
    link: str
    date: str
    summary: str
    source: str
    category: str = ""
    keywords: List[str] = field(default_factory=list)
    content_hash: str = ""


class SecurityNewsCollector:
    """
    資安新聞收集器
    
    功能：
    - 從多個 RSS 來源收集新聞
    - 自動分類（惡意程式、漏洞、資料外洩等）
    - 關鍵字擷取
    - 搜尋與篩選
    - 匯出 Excel/CSV/JSON
    """
    
    # 預設新聞來源
    DEFAULT_SOURCES = {
        # 台灣來源
        'iThome 資安': 'https://www.ithome.com.tw/rss/security',
        'TWCERT/CC': 'https://www.twcert.org.tw/rss',
        
        # 國際來源
        'The Hacker News': 'https://feeds.feedburner.com/TheHackersNews',
        'Krebs on Security': 'https://krebsonsecurity.com/feed/',
        'BleepingComputer': 'https://www.bleepingcomputer.com/feed/',
        'Dark Reading': 'https://www.darkreading.com/rss.xml',
        'SecurityWeek': 'https://feeds.feedburner.com/securityweek',
        'Threatpost': 'https://threatpost.com/feed/',
        'HackRead': 'https://www.hackread.com/feed/',
        'Sophos News': 'https://news.sophos.com/en-us/feed/',
    }
    
    # 分類關鍵字
    CATEGORY_KEYWORDS = {
        '惡意程式/Malware': ['malware', 'ransomware', 'trojan', 'virus', 'worm', 'botnet', 
                           '惡意程式', '勒索軟體', '木馬', '病毒', '蠕蟲', '殭屍網路'],
        '漏洞/Vulnerability': ['vulnerability', 'cve', 'exploit', 'zero-day', '0day', 'patch',
                              '漏洞', '弱點', '修補', '零日'],
        '資料外洩/Data Breach': ['breach', 'leak', 'exposed', 'stolen', 'dump',
                               '外洩', '洩漏', '竊取', '曝光'],
        '駭客攻擊/Hacking': ['hack', 'attack', 'intrusion', 'compromise', 'apt',
                           '駭客', '攻擊', '入侵', '滲透'],
        '網路釣魚/Phishing': ['phishing', 'scam', 'fraud', 'social engineering',
                            '釣魚', '詐騙', '社交工程'],
        '供應鏈/Supply Chain': ['supply chain', 'software supply', 'dependency',
                               '供應鏈', '軟體供應'],
        '雲端安全/Cloud Security': ['cloud', 'aws', 'azure', 'gcp', 'kubernetes', 'container',
                                  '雲端', '容器'],
        '物聯網/IoT': ['iot', 'smart device', 'embedded', 'firmware',
                      '物聯網', '智慧裝置', '韌體'],
        '政策法規/Policy': ['regulation', 'compliance', 'gdpr', 'policy', 'law',
                          '法規', '合規', '政策', '法律'],
    }
    
    # 重要關鍵字（用於擷取）
    IMPORTANT_KEYWORDS = [
        'CVE', 'APT', 'zero-day', 'ransomware', 'malware', 'phishing',
        'vulnerability', 'exploit', 'breach', 'attack', 'hack',
        'Microsoft', 'Google', 'Apple', 'Linux', 'Windows', 'Android', 'iOS',
        'FBI', 'NSA', 'CISA', 'Mandiant', 'CrowdStrike',
        '漏洞', '駭客', '攻擊', '勒索', '惡意程式', '資安'
    ]
    
    def __init__(self, custom_sources: Dict[str, str] = None, sources_file: str = None):
        """
        初始化收集器
        
        Args:
            custom_sources: 自訂來源字典 {名稱: RSS URL}
            sources_file: 來源設定 JSON 檔案路徑
        """
        self.sources = self.DEFAULT_SOURCES.copy()
        
        # 從檔案載入來源
        if sources_file and os.path.exists(sources_file):
            self.load_sources_from_file(sources_file)
        
        # 合併自訂來源
        if custom_sources:
            self.sources.update(custom_sources)
        
        self.news_items: List[NewsItem] = []
        self.df: Optional[pd.DataFrame] = None
        self.seen_hashes: set = set()
    
    def load_sources_from_file(self, filepath: str) -> None:
        """從 JSON 檔案載入來源設定"""
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                loaded_sources = json.load(f)
                self.sources.update(loaded_sources)
                print(f"✅ 已從 {filepath} 載入 {len(loaded_sources)} 個來源")
        except Exception as e:
            print(f"⚠️ 載入來源檔案失敗: {e}")
    
    def save_sources_to_file(self, filepath: str) -> None:
        """儲存來源設定到 JSON 檔案"""
        try:
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(self.sources, f, ensure_ascii=False, indent=2)
            print(f"✅ 已儲存 {len(self.sources)} 個來源到 {filepath}")
        except Exception as e:
            print(f"⚠️ 儲存來源檔案失敗: {e}")
    
    def add_source(self, name: str, url: str) -> None:
        """新增來源"""
        self.sources[name] = url
        print(f"✅ 已新增來源: {name}")
    
    def remove_source(self, name: str) -> bool:
        """移除來源"""
        if name in self.sources:
            del self.sources[name]
            print(f"✅ 已移除來源: {name}")
            return True
        print(f"⚠️ 找不到來源: {name}")
        return False
    
    def list_sources(self) -> None:
        """列出所有來源"""
        print(f"\n📰 新聞來源清單 (共 {len(self.sources)} 個)")
        print("-" * 60)
        for name, url in self.sources.items():
            is_default = "📌" if name in self.DEFAULT_SOURCES else "➕"
            print(f"{is_default} {name}")
            print(f"   {url}")
        print("-" * 60)
    
    def _generate_hash(self, content: str) -> str:
        """產生內容雜湊值（用於去重）"""
        return hashlib.md5(content.encode('utf-8')).hexdigest()
    
    def _clean_html(self, html_content: str) -> str:
        """清理 HTML 標籤，保留純文字"""
        if not html_content:
            return ""
        soup = BeautifulSoup(html_content, 'lxml')
        text = soup.get_text(separator=' ', strip=True)
        # 移除多餘空白
        text = re.sub(r'\s+', ' ', text)
        return text.strip()
    
    def _extract_summary(self, entry: dict, max_length: int = 300) -> str:
        """擷取摘要"""
        # 嘗試不同的摘要欄位
        summary = ""
        
        if hasattr(entry, 'summary'):
            summary = entry.summary
        elif hasattr(entry, 'description'):
            summary = entry.description
        elif hasattr(entry, 'content') and entry.content:
            summary = entry.content[0].get('value', '')
        
        # 清理 HTML
        summary = self._clean_html(summary)
        
        # 截斷到最大長度
        if len(summary) > max_length:
            summary = summary[:max_length].rsplit(' ', 1)[0] + '...'
        
        return summary
    
    def _parse_date(self, entry: dict) -> str:
        """解析日期"""
        try:
            if hasattr(entry, 'published_parsed') and entry.published_parsed:
                dt = datetime(*entry.published_parsed[:6])
                return dt.strftime('%Y-%m-%d %H:%M')
            elif hasattr(entry, 'updated_parsed') and entry.updated_parsed:
                dt = datetime(*entry.updated_parsed[:6])
                return dt.strftime('%Y-%m-%d %H:%M')
            elif hasattr(entry, 'published'):
                return entry.published[:16]
        except:
            pass
        return datetime.now().strftime('%Y-%m-%d %H:%M')
    
    def _categorize(self, title: str, summary: str) -> str:
        """自動分類新聞"""
        content = (title + ' ' + summary).lower()
        
        for category, keywords in self.CATEGORY_KEYWORDS.items():
            for keyword in keywords:
                if keyword.lower() in content:
                    return category
        
        return '其他/Other'
    
    def _extract_keywords(self, title: str, summary: str) -> List[str]:
        """擷取關鍵字"""
        content = title + ' ' + summary
        keywords = []
        
        # 找 CVE 編號
        cve_pattern = r'CVE-\d{4}-\d+'
        cves = re.findall(cve_pattern, content, re.IGNORECASE)
        keywords.extend([cve.upper() for cve in cves])
        
        # 找重要關鍵字
        for kw in self.IMPORTANT_KEYWORDS:
            if kw.lower() in content.lower() and kw not in keywords:
                keywords.append(kw)
        
        return keywords[:10]  # 最多 10 個關鍵字
    
    def _fetch_feed(self, source_name: str, url: str) -> List[NewsItem]:
        """抓取單一 RSS Feed"""
        items = []
        try:
            feed = feedparser.parse(url)
            
            for entry in feed.entries[:20]:  # 每個來源最多 20 則
                title = entry.get('title', '').strip()
                link = entry.get('link', '').strip()
                
                if not title or not link:
                    continue
                
                # 去重檢查
                content_hash = self._generate_hash(title + link)
                if content_hash in self.seen_hashes:
                    continue
                self.seen_hashes.add(content_hash)
                
                # 擷取資訊
                summary = self._extract_summary(entry)
                date = self._parse_date(entry)
                category = self._categorize(title, summary)
                keywords = self._extract_keywords(title, summary)
                
                item = NewsItem(
                    title=title,
                    link=link,
                    date=date,
                    summary=summary,
                    source=source_name,
                    category=category,
                    keywords=keywords,
                    content_hash=content_hash
                )
                items.append(item)
            
            print(f"✅ {source_name}: {len(items)} 則新聞")
            
        except Exception as e:
            print(f"❌ {source_name}: 抓取失敗 - {str(e)[:50]}")
        
        return items
    
    def collect(self, max_workers: int = 5) -> pd.DataFrame:
        """
        收集所有來源的新聞
        
        Args:
            max_workers: 並行抓取的最大執行緒數
            
        Returns:
            包含所有新聞的 DataFrame
        """
        print(f"\n🔍 開始收集 {len(self.sources)} 個新聞來源...")
        print("-" * 50)
        
        self.news_items = []
        self.seen_hashes = set()
        
        # 並行抓取
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(self._fetch_feed, name, url): name 
                for name, url in self.sources.items()
            }
            
            for future in as_completed(futures):
                items = future.result()
                self.news_items.extend(items)
        
        print("-" * 50)
        print(f"📊 共收集 {len(self.news_items)} 則不重複新聞")
        
        # 轉換為 DataFrame
        if self.news_items:
            self.df = pd.DataFrame([asdict(item) for item in self.news_items])
            self.df['keywords_str'] = self.df['keywords'].apply(
                lambda x: ', '.join(x) if x else ''
            )
            # 依日期排序
            self.df = self.df.sort_values('date', ascending=False).reset_index(drop=True)
        else:
            self.df = pd.DataFrame()
        
        return self.df
    
    def search(self, query: str) -> pd.DataFrame:
        """
        搜尋新聞
        
        Args:
            query: 搜尋關鍵字
            
        Returns:
            符合的新聞 DataFrame
        """
        if self.df is None or self.df.empty:
            return pd.DataFrame()
        
        query_lower = query.lower()
        mask = (
            self.df['title'].str.lower().str.contains(query_lower, na=False) |
            self.df['summary'].str.lower().str.contains(query_lower, na=False) |
            self.df['keywords_str'].str.lower().str.contains(query_lower, na=False)
        )
        return self.df[mask].reset_index(drop=True)
    
    def filter_by_category(self, category: str) -> pd.DataFrame:
        """依分類篩選"""
        if self.df is None or self.df.empty:
            return pd.DataFrame()
        
        mask = self.df['category'].str.contains(category, case=False, na=False)
        return self.df[mask].reset_index(drop=True)
    
    def filter_by_source(self, source: str) -> pd.DataFrame:
        """依來源篩選"""
        if self.df is None or self.df.empty:
            return pd.DataFrame()
        
        mask = self.df['source'].str.contains(source, case=False, na=False)
        return self.df[mask].reset_index(drop=True)
    
    def filter_by_date(self, start_date: str = None, end_date: str = None) -> pd.DataFrame:
        """依日期範圍篩選"""
        if self.df is None or self.df.empty:
            return pd.DataFrame()
        
        result = self.df.copy()
        
        if start_date:
            result = result[result['date'] >= start_date]
        if end_date:
            result = result[result['date'] <= end_date]
        
        return result.reset_index(drop=True)
    
    def get_summary_stats(self) -> Dict:
        """取得統計摘要"""
        if self.df is None or self.df.empty:
            return {}
        
        return {
            'total_news': len(self.df),
            'sources': self.df['source'].value_counts().to_dict(),
            'categories': self.df['category'].value_counts().to_dict(),
            'date_range': {
                'earliest': self.df['date'].min(),
                'latest': self.df['date'].max()
            }
        }
    
    def to_excel(self, filepath: str = 'security_news_report.xlsx') -> None:
        """
        匯出 Excel 報表（含格式化）
        
        Args:
            filepath: 輸出檔案路徑
        """
        if self.df is None or self.df.empty:
            print("⚠️ 沒有資料可匯出")
            return
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        wb = Workbook()
        
        # ===== 工作表 1: 新聞清單 =====
        ws1 = wb.active
        ws1.title = "資安新聞"
        
        # 標題列
        headers = ['序號', '日期', '來源', '分類', '標題', '摘要', '關鍵字', '連結']
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 分類顏色
        category_colors = {
            '惡意程式': 'FFE0E0',
            '漏洞': 'FFF3E0',
            '資料外洩': 'F3E5F5',
            '駭客攻擊': 'E3F2FD',
            '網路釣魚': 'FFF8E1',
            '供應鏈': 'E8F5E9',
            '雲端安全': 'E0F7FA',
            '物聯網': 'FBE9E7',
            '政策法規': 'ECEFF1',
        }
        
        # 資料列
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for idx, row in self.df.iterrows():
            row_num = idx + 2
            
            ws1.cell(row=row_num, column=1, value=idx + 1)
            ws1.cell(row=row_num, column=2, value=row['date'])
            ws1.cell(row=row_num, column=3, value=row['source'])
            
            # 分類（帶顏色）
            cat_cell = ws1.cell(row=row_num, column=4, value=row['category'])
            for cat_key, color in category_colors.items():
                if cat_key in row['category']:
                    cat_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    break
            
            ws1.cell(row=row_num, column=5, value=row['title'])
            ws1.cell(row=row_num, column=6, value=row['summary'])
            ws1.cell(row=row_num, column=7, value=row['keywords_str'])
            
            # 連結
            link_cell = ws1.cell(row=row_num, column=8, value=row['link'])
            link_cell.hyperlink = row['link']
            link_cell.font = Font(color="0563C1", underline="single")
            
            # 套用框線
            for col in range(1, 9):
                ws1.cell(row=row_num, column=col).border = thin_border
                ws1.cell(row=row_num, column=col).alignment = Alignment(
                    wrap_text=True, vertical='top'
                )
        
        # 調整欄寬
        ws1.column_dimensions['A'].width = 6
        ws1.column_dimensions['B'].width = 16
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 20
        ws1.column_dimensions['E'].width = 50
        ws1.column_dimensions['F'].width = 60
        ws1.column_dimensions['G'].width = 30
        ws1.column_dimensions['H'].width = 40
        
        # 凍結首列
        ws1.freeze_panes = 'A2'
        
        # ===== 工作表 2: 統計分析 =====
        ws2 = wb.create_sheet(title="統計分析")
        
        stats = self.get_summary_stats()
        
        # 標題
        ws2.cell(row=1, column=1, value="📊 資安新聞統計報告")
        ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws2.merge_cells('A1:C1')
        
        # 總覽
        ws2.cell(row=3, column=1, value="總覽")
        ws2.cell(row=3, column=1).font = Font(bold=True)
        ws2.cell(row=4, column=1, value="新聞總數")
        ws2.cell(row=4, column=2, value=stats.get('total_news', 0))
        ws2.cell(row=5, column=1, value="最早日期")
        ws2.cell(row=5, column=2, value=stats.get('date_range', {}).get('earliest', ''))
        ws2.cell(row=6, column=1, value="最新日期")
        ws2.cell(row=6, column=2, value=stats.get('date_range', {}).get('latest', ''))
        
        # 來源統計
        ws2.cell(row=8, column=1, value="來源分布")
        ws2.cell(row=8, column=1).font = Font(bold=True)
        row_num = 9
        for source, count in stats.get('sources', {}).items():
            ws2.cell(row=row_num, column=1, value=source)
            ws2.cell(row=row_num, column=2, value=count)
            row_num += 1
        
        # 分類統計
        row_num += 1
        ws2.cell(row=row_num, column=1, value="分類分布")
        ws2.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        for category, count in stats.get('categories', {}).items():
            ws2.cell(row=row_num, column=1, value=category)
            ws2.cell(row=row_num, column=2, value=count)
            row_num += 1
        
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 15
        
        # 儲存
        wb.save(filepath)
        print(f"📊 Excel 報表已匯出至 {filepath}")
    
    def to_csv(self, filepath: str = 'security_news.csv') -> None:
        """匯出 CSV"""
        if self.df is None or self.df.empty:
            print("⚠️ 沒有資料可匯出")
            return
        
        export_df = self.df[['date', 'source', 'category', 'title', 'summary', 'keywords_str', 'link']]
        export_df.to_csv(filepath, index=False, encoding='utf-8-sig')
        print(f"💾 已匯出至 {filepath}")
    
    def to_json(self, filepath: str = 'security_news.json') -> None:
        """匯出 JSON"""
        if self.df is None or self.df.empty:
            print("⚠️ 沒有資料可匯出")
            return
        
        data = {
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'total_count': len(self.df),
            'sources': list(self.df['source'].unique()),
            'categories': list(self.df['category'].unique()),
            'news': self.df.to_dict(orient='records')
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        print(f"💾 已匯出至 {filepath}")


def load_demo_data() -> List[Dict]:
    """載入展示資料（當無法連線時使用）"""
    return [
        {
            "title": "微軟發布緊急更新修補 WSUS 重大漏洞 CVE-2025-59287",
            "link": "https://example.com/news/1",
            "date": "2025-11-24 14:30",
            "summary": "微軟緊急發布更新修補 Windows Server Update Services (WSUS) 中的重大遠端程式碼執行漏洞。該漏洞允許未經驗證的攻擊者在目標系統上執行任意程式碼。",
            "source": "The Hacker News",
            "category": "漏洞/Vulnerability",
            "keywords": ["CVE-2025-59287", "Microsoft", "WSUS", "vulnerability"]
        },
        {
            "title": "CrazyHunter 勒索軟體針對台灣金融業發動攻擊",
            "link": "https://example.com/news/2",
            "date": "2025-11-23 10:15",
            "summary": "資安研究人員發現新型勒索軟體 CrazyHunter 正針對台灣金融機構發動攻擊，該惡意程式使用雙重勒索策略，除加密檔案外還威脅公開敏感資料。",
            "source": "iThome 資安",
            "category": "惡意程式/Malware",
            "keywords": ["ransomware", "CrazyHunter", "勒索軟體", "金融業"]
        },
        {
            "title": "研究人員發現 AI 模型可被利用產生惡意程式碼",
            "link": "https://example.com/news/3",
            "date": "2025-11-22 16:45",
            "summary": "資安研究團隊發表報告指出，大型語言模型可能被惡意行為者利用來產生惡意程式碼，呼籲 AI 開發商加強安全防護措施。",
            "source": "SecurityWeek",
            "category": "其他/Other",
            "keywords": ["AI", "malware", "研究"]
        },
        {
            "title": "Oracle 資料庫伺服器發現高風險 SQL 注入漏洞",
            "link": "https://example.com/news/4",
            "date": "2025-11-21 09:20",
            "summary": "Oracle 發布安全公告，針對其資料庫伺服器產品中的 SQL 注入漏洞 CVE-2025-61757 提供修補程式，該漏洞 CVSS 評分高達 9.8。",
            "source": "Dark Reading",
            "category": "漏洞/Vulnerability",
            "keywords": ["CVE-2025-61757", "Oracle", "SQL injection", "vulnerability"]
        },
        {
            "title": "中國駭客組織 APT24 針對東南亞政府機關發動攻擊",
            "link": "https://example.com/news/5",
            "date": "2025-11-20 11:30",
            "summary": "威脅情報公司揭露中國國家級駭客組織 APT24 近期針對東南亞多國政府機關發動網路間諜活動，竊取敏感外交文件。",
            "source": "The Hacker News",
            "category": "駭客攻擊/Hacking",
            "keywords": ["APT24", "China", "駭客", "政府"]
        },
        {
            "title": "台灣企業遭供應鏈攻擊，數百家公司受影響",
            "link": "https://example.com/news/6",
            "date": "2025-11-19 15:00",
            "summary": "調查局資安工作站警告，駭客透過入侵本土軟體供應商，植入惡意程式後門，影響使用該軟體的數百家台灣企業。",
            "source": "TWCERT/CC",
            "category": "供應鏈/Supply Chain",
            "keywords": ["supply chain", "供應鏈", "台灣", "malware"]
        },
        {
            "title": "AWS S3 儲存桶配置錯誤導致百萬用戶資料外洩",
            "link": "https://example.com/news/7",
            "date": "2025-11-18 13:45",
            "summary": "某知名 SaaS 平台因 AWS S3 儲存桶配置不當，導致超過 100 萬用戶個資外洩，包括姓名、電子郵件及加密密碼。",
            "source": "BleepingComputer",
            "category": "資料外洩/Data Breach",
            "keywords": ["AWS", "S3", "data breach", "雲端安全"]
        },
        {
            "title": "新型網路釣魚攻擊偽裝成 Microsoft 365 登入頁面",
            "link": "https://example.com/news/8",
            "date": "2025-11-17 10:00",
            "summary": "資安公司發現新型釣魚攻擊活動，駭客建立高度仿真的 Microsoft 365 登入頁面，已有多家企業員工受害。",
            "source": "Krebs on Security",
            "category": "網路釣魚/Phishing",
            "keywords": ["phishing", "Microsoft 365", "釣魚"]
        },
        {
            "title": "智慧家電漏洞允許駭客遠端控制設備",
            "link": "https://example.com/news/9",
            "date": "2025-11-16 14:20",
            "summary": "研究人員在多款智慧家電中發現重大安全漏洞，攻擊者可利用這些漏洞遠端控制設備，甚至監聽用戶對話。",
            "source": "HackRead",
            "category": "物聯網/IoT",
            "keywords": ["IoT", "物聯網", "smart home", "vulnerability"]
        },
        {
            "title": "歐盟通過新網路安全法規強化關鍵基礎設施保護",
            "link": "https://example.com/news/10",
            "date": "2025-11-15 09:00",
            "summary": "歐盟議會通過新版網路安全法規，要求關鍵基礎設施營運商強化資安措施，違者將面臨高額罰款。",
            "source": "SecurityWeek",
            "category": "政策法規/Policy",
            "keywords": ["EU", "法規", "policy", "關鍵基礎設施"]
        },
        {
            "title": "LockBit 勒索軟體集團宣稱入侵多國醫療機構",
            "link": "https://example.com/news/11",
            "date": "2025-11-14 16:30",
            "summary": "惡名昭彰的 LockBit 勒索軟體集團在暗網公布多家醫療機構資料，要求支付贖金否則將公開病患個資。",
            "source": "TWCERT/CC",
            "category": "惡意程式/Malware",
            "keywords": ["LockBit", "ransomware", "勒索軟體", "醫療"]
        },
        {
            "title": "Grafana 發布重大安全更新修補認證繞過漏洞",
            "link": "https://example.com/news/12",
            "date": "2025-11-14 08:45",
            "summary": "開源監控平台 Grafana 發布緊急安全更新，修補 CVE-2025-41115 認證繞過漏洞，建議用戶立即升級。",
            "source": "iThome 資安",
            "category": "漏洞/Vulnerability",
            "keywords": ["CVE-2025-41115", "Grafana", "vulnerability", "認證繞過"]
        }
    ]


# 主程式（命令列使用）
if __name__ == '__main__':
    print("=" * 60)
    print("🛡️  資安新聞收集器")
    print("=" * 60)
    
    collector = SecurityNewsCollector()
    df = collector.collect()
    
    if df.empty:
        print("\n⚠️ 無法連線至新聞來源，載入展示資料...")
        demo_data = load_demo_data()
        
        for item in demo_data:
            news_item = NewsItem(
                title=item['title'],
                link=item['link'],
                date=item['date'],
                summary=item['summary'],
                source=item['source'],
                category=item['category'],
                keywords=item['keywords'],
                content_hash=collector._generate_hash(item['title'] + item['link'])
            )
            collector.news_items.append(news_item)
        
        collector.df = pd.DataFrame([asdict(item) for item in collector.news_items])
        collector.df['keywords_str'] = collector.df['keywords'].apply(
            lambda x: ', '.join(x) if x else ''
        )
        collector.df = collector.df.sort_values('date', ascending=False).reset_index(drop=True)
        print(f"📊 已載入 {len(collector.df)} 則展示新聞")
    
    # 匯出檔案
    os.makedirs('site', exist_ok=True)
    collector.to_json('site/data.json')
    collector.to_excel('site/security_news_report.xlsx')
    collector.to_csv('site/security_news.csv')
    
    # 顯示統計
    stats = collector.get_summary_stats()
    print(f"\n📈 統計摘要:")
    print(f"   新聞總數: {stats.get('total_news', 0)}")
    print(f"   來源數: {len(stats.get('sources', {}))}")
    print(f"   分類數: {len(stats.get('categories', {}))}")
